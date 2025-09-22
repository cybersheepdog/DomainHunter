import logging
import dnstwist
import whois
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import date
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from configparser import ConfigParser

# Configure logging
logging.basicConfig(
        filename="domainhunter.log",
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
)


# Reads in domains to be monitored from a text file and puts them in a list.  Domains are 1 per line.
domain_list = open("monitored_domains.txt", "r")
domain_list = domain_list.read()
domain_list = domain_list.split("\n")
domain_list = [domain for domain in domain_list if domain]

# Declare abused TLD domains for permutations
abused_dict = "abused_tlds.dict"

# Get email config from config file
config_object = ConfigParser()
try:
    config_object.read("config.ini")
except:
    logging.info('Error with config.ini')
else:
    email = config_object["EMAIL"]
    if email['password']:
        password = email['password']
    else:
        logging.info('Email password not configured in config.ini.  Please do so before proceeding.')
        exit()
    if email['receiver_email']:
        receiver_email = email['receiver_email']
        receiver_email = receiver_email.split(",")
    else:
        logging.info('Receiver Email(s) not configured in config.ini.  Please do so before preoceeding.')
        exit()
    if email['sender_email']:
        sender_email = email['sender_email']
    else:
        logging.info('Sender Email not configured in config.ini.  Please do so before preoceeding.')
        exit()
    
# Declare dynamic variabls for sending emails
body = ""
subject = ""

def append_new_domains(registered_domains, new_domains, file_name, rows, receiver_email, today):
    wb = load_workbook(file_name)
    ws = wb.active
    for dom in registered_domains:
        if dom["domain"] in new_domains:
            ws['A' + str(rows)] = today
            ws['B' + str(rows)] = dom["domain"]
            ws['C' + str(rows)] = dom["fuzzer"]
            if type(dom["Created Date"]) == list:
                if dom["Created Date"][0].date() == dom["Created Date"][1].date():
                    ws['D' + str(rows)] = str(dom["Created Date"][0].date())
                else:
                    ws['D' + str(rows)] = str(dom["Created Date"][0].date())
                    ws['E' + str(rows)] = str(dom["Created Date"][1].date())
            else:
                try:
                    dom["Created Date"].date()
                except:
                    pass
                else:
                    ws['D' + str(rows)] = str(dom["Created Date"].date())
            try:
                dom["Name"]
            except:
                ws['H' + str(rows)] = ""
            else:
                if type(dom["Name"]) == list:
                    for name in dom["Name"]:
                        if name == 'REDACTED FOR PRIVACY':
                            pass
                        else:
                            ws['H' + str(rows)] = name
                else:
                    ws['H' + str(rows)] = dom["Name"]
            try:
                dom["Org"]
            except:
                ws['I' + str(rows)] = ""
            else:
                ws['I' + str(rows)] = dom["Org"]
            try:
                dom["phash"]
            except:
                ws['J' + str(rows)] = ""
            else:
                ws['J' + str(rows)] = dom["phash"]
            try:
                dom["dns_ns"]
            except:
                ws['K' + str(rows)] = ""
            else:
                ws['K' + str(rows)] = dom["dns_ns"][0]
            try:
                dom["dns_a"]
            except:
                ws['L' + str(rows)] = ""
            else:
                ws['L' + str(rows)] = dom["dns_a"][0]
            try:
                dom["dns_mx"]
            except:
                ws['M' + str(rows)] = ""
            else:
                ws['M' + str(rows)] = dom["dns_mx"][0]
            try:
                dom["Emails"]
            except:
                ws['N' + str(rows)] = ""
            else:
                if type(dom["Emails"]) == list:
                    ws['N' + str(rows)] = dom["Emails"][0]
                    ws['O' + str(rows)] = dom["Emails"][1]                
                elif type(dom["Emails"]) == str:
                    ws['N' + str(rows)] = dom["Emails"]
                else:
                    pass
            rows = rows + 1                         
            
            wb.save(file_name)
            wb.close()


            # Fill in variables for new alert email
            client_name = client[0].title()
            subject = "New domain alert for %s." % client_name
            text = """\
                   Hello,\n
                   Here is the new domain alert for %s\n.
                   New Domain: %s\n
                   Created Date: %s\n
                   Type: %s
                   """ % (client_name, dom["domain"], dom["Created Date"], dom["fuzzer"])
            html = """\
            <html>
              <body>
                <p>Hello,<br>
                   Here is the new domain alert for %s.<br>
                   New Domain: %s<br>
                   Created Date: %s<br>
                  Type: %s<br>
                </p>
              </body>
            </html>
            """ % (client_name, dom["domain"], dom["Created Date"], dom["fuzzer"])
            
            send_new_domain_alert_email(subject, text, html, sender_email,receiver_email, file_name, password)
            logging.info('Email sent for {} {}'.format(client_name,dom["domain"]))
        else:
            pass

def create_fill_initial_excel_for_domain(file_name, registered_domains):
    # Create new Excel Document
    wb = Workbook()
    # Get active sheet
    ws = wb.active

    # Create Column Names
    ws['A1'] = "Added Date"
    ws['A1'].font = Font(bold=True)
    ws['B1'] = "Domain"
    ws['B1'].font = Font(bold=True)
    ws['C1'] = "Permutation"
    ws['C1'].font = Font(bold=True)
    ws['D1'] = "Date Created 1"
    ws['D1'].font = Font(bold=True)
    ws['E1'] = "Date Created 2"
    ws['E1'].font = Font(bold=True)
    ws['F1'] = "Last Updated 1"
    ws['F1'].font = Font(bold=True)
    ws['G1'] = "Last Updated 2"
    ws['G1'].font = Font(bold=True)
    ws['H1'] = "Registrant Name"
    ws['H1'].font = Font(bold=True)
    ws['I1'] = "Organization"
    ws['I1'].font = Font(bold=True)
    ws['J1'] = "PHash"
    ws['J1'].font = Font(bold=True)
    ws['K1'] = "Name Server"
    ws['K1'].font = Font(bold=True)
    ws['L1'] = "IP"
    ws['L1'].font = Font(bold=True)
    ws['M1'] = "Mail Server"
    ws['M1'].font = Font(bold=True)
    ws['N1'] = "Registered Email 1"
    ws['N1'].font = Font(bold=True)
    ws['O1'] = "Registered Email 2"
    ws['O1'].font = Font(bold=True)
    # Turn on auto-filter
    ws.auto_filter.ref = ws.dimensions
    count = 2
    for dom in registered_domains:
        ws['A' + str(count)] = today
        ws['B' + str(count)] = dom["domain"]
        ws['C' + str(count)] = dom["fuzzer"]
        if type(dom["Created Date"]) == list:
            if dom["Created Date"][0].date() == dom["Created Date"][1].date():
                ws['D' + str(count)] = str(dom["Created Date"][0].date())
            else:
                ws['D' + str(count)] = str(dom["Created Date"][0].date())
                ws['E' + str(count)] = str(dom["Created Date"][1].date())
        else:
            try:
                dom["Created Date"].date()
            except:
                pass
            else:
                ws['D' + str(count)] = str(dom["Created Date"].date())
        try:
            dom["Name"]
        except:
            ws['H' + str(count)] = ""
        else:
            if type(dom["Name"]) == list:
                for name in dom["Name"]:
                    if name == 'REDACTED FOR PRIVACY':
                        pass
                    else:
                        ws['H' + str(count)] = name
            else:
                ws['H' + str(count)] = dom["Name"]
        try:
            dom["Org"]
        except:
            ws['I' + str(count)] = ""
        else:
            ws['I' + str(count)] = dom["Org"]
        try:
            dom["phash"]
        except:
            ws['J' + str(count)] = ""
        else:
            ws['J' + str(count)] = dom["phash"]
        try:
            dom["dns_ns"]
        except:
            ws['K' + str(count)] = ""
        else:
            ws['K' + str(count)] = dom["dns_ns"][0]
        try:
            dom["dns_a"]
        except:
            ws['L' + str(count)] = ""
        else:
            ws['L' + str(count)] = dom["dns_a"][0]
        try:
            dom["dns_mx"]
        except:
            ws['M' + str(count)] = ""
        else:
            ws['M' + str(count)] = dom["dns_mx"][0]
        try:
            dom["Emails"]
        except:
            ws['N' + str(count)] = ""
        else:
            if type(dom["Emails"]) == list:
                ws['N' + str(count)] = dom["Emails"][0]
                ws['O' + str(count)] = dom["Emails"][1]                
            elif type(dom["Emails"]) == str:
                ws['N' + str(count)] = dom["Emails"]
            else:
                pass
                        
        count = count + 1
    # Save and close Excel Document
    wb.save(file_name)
    wb.close()
    
def get_new_domains(file_name, registered_domains, existing_domains, new_domains, rows):
    wb = load_workbook(file_name)
    ws = wb.active
    for col in ws['B']:
        rows = rows + 1
        if col.value == "Domain":
            pass
        else:
            existing_domains.append(col.value)

    for line in registered_domains:
        if line["domain"] in existing_domains:
            pass
        else:
            new_domains.append(line["domain"])
            
    wb.save(file_name)
    wb.close()
    print("rows: " + str(rows))
    return rows
    
def get_registered_permutations(monitored_domain, registered_domains):
    logging.info('Start of looking for permutations for (%s)' % (monitored_domain))
    registered_domains = dnstwist.run(domain=monitored_domain, registered=True, phash=True, format='null')
    logging.info('End  of looking for permutations for (%s)' % (monitored_domain)) 
    logging.info('Start of whois for registered domains for (%s)' % (monitored_domain))
    
    for domain in registered_domains:
        try:
            w = whois.whois(domain["domain"])
        except Exception:
            pass
        else:
            domain["Created Date"] = w.creation_date
            domain["Org"] = w.org
            domain["Name"] = w.name
            domain["Emails"] = w.emails
            domain["Updated Date"] = w.updated_date
            domain["Expiration Date"] = w.expiration_date
            domain["Status"] = w.status

    logging. info('End of whois for registered domains for (%s)' % (monitored_domain))
        
    return registered_domains
            
def send_email_with_attachment(subject, body, sender_email,receiver_email, file_name, password):
    to = ",".join(receiver_email)

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = "TH Domain Alerts"
    message["To"] = to
    message["Subject"] = subject
    message["Bcc"] = ''#receiver_email  # Recommended for mass emails

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    #file_name = ""  # In same directory as script

    # Open PDF file in binary mode
    with open(file_name, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {file_name}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, text)

    logging.info('Initial registered domain email sent.')
        
def send_new_domain_alert_email(subject, text, html, sender_email, receiver_email, file_name, password):
    to = ",".join(receiver_email)
    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = "TH Domain Alerts"
    message["To"] = to

    # Turn these into plain/html MIMEText objects
    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part1)
    message.attach(part2)
    
    #file_name = ""  # In same directory as script

    # Open PDF file in binary mode
    with open(file_name, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {file_name}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()


    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())

    logging.info('New domain email sent.')

logging.info('Start of program')


for monitored_domain in domain_list:
    client = monitored_domain.split(".")
    file_name = client[0] + ".xlsx"
    registered = []
    registered_domains = []
    today = date.today()

    registered = get_registered_permutations(monitored_domain, registered_domains)
    registered_domains = registered

    try:
        wb = load_workbook(file_name)
    except:
        create_fill_initial_excel_for_domain(file_name, registered_domains)
        client_name = client[0].title()
        subject = "TH: Initial similar domain report for %s." % client_name
        body = "Hello,\nHere is the Excel document containing all similarly registered domains for %s for %s." % (monitored_domain, client_name)
        send_email_with_attachment(subject, body, sender_email, receiver_email, file_name, password)
    else:
        existing_domains = []
        new_domains = []
        rows = 1
        rows = get_new_domains(file_name, registered_domains, existing_domains, new_domains, rows)
        append_new_domains(registered_domains, new_domains, file_name, rows, receiver_email, today)

logging.info('End of program')
